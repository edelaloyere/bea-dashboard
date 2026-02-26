"""
Script de mise à jour du dashboard Béa
========================================
Ce script lit le fichier Excel et génère public/data/retombees.json

Usage :
    python -X utf8 update_dashboard.py

Prérequis :
    python -m pip install openpyxl

Structure Excel attendue (Feuil1) :
    Col 1: Média | Col 2: Type | Col 3: Date | Col 4: Titre | Col 5: URL
"""

import openpyxl
import json
import os
from datetime import datetime

# ─── CONFIGURATION ────────────────────────────────────────────────────────────

EXCEL_FILE = "Bilan des retombées à jour (1).xlsx"
OUTPUT_JSON = os.path.join("public", "data", "retombees.json")

LOGOS_MEDIAS = {
    "BFM Business":                       "/logos/bfm.png",
    "BFMTV":                              "/logos/bfm.png",
    "BFMTV ":                             "/logos/bfm.png",
    "Capital":                            "/logos/capital.png",
    "Les Echos":                          "/logos/les-echos.png",
    "Les Echos ":                         "/logos/les-echos.png",
    "Les Echos Week-end":                 "/logos/les-echos.png",
    "Les Echos Week-end ":                "/logos/les-echos.png",
    "Le Figaro":                          "/logos/le-figaro.png",
    "Le Particulier (Groupe Figaro)":     "/logos/le-particulier.png",
    "Le Parisien / Aujourd'hui en France":"/logos/le-parisien.png",
    "La Parisien / Aujourd'hui en France":"/logos/le-parisien.png",
    "Ouest-France":                       "/logos/ouest-france.png",
    "Programme TV Ouest-France":          "/logos/ouest-france.png",
    "Mieux Vivre Votre Argent":           "/logos/investir.png",
    "Sud Radio":                          "/logos/sud-radio.png",
    "Europe 1":                           "/logos/europe1.png",
    "RMC":                                "/logos/rmc.png",
    "France 2":                           "/logos/france-2.png",
    "TF1":                                "/logos/tf1.png",
    "MSN":                                "/logos/msn.png",
    "Maddyness":                          "/logos/maddyness.png",
    "Ideal Investisseur":                 "/logos/ideal-investisseur.png",
    "My Sweet Immo":                      "/logos/mysweetimmo.png",
    "My Sweet Immo (YouTube)":            "/logos/mysweetimmo.png",
    "Compte My Sweet Immo":               "/logos/mysweetimmo.png",
}

# Médias considérés comme "clés" (tier-1 national)
# NB : Les Echos Week-end, Programme TV Ouest-France, Sud Radio etc. ne comptent pas séparément
MEDIAS_CLES_REF = {
    "BFM Business", "BFMTV", "BFMTV ",
    "Capital",
    "France 2", "TF1",
    "Le Figaro",
    "Le Parisien / Aujourd'hui en France", "La Parisien / Aujourd'hui en France",
    "Les Echos", "Les Echos ",
    "Mieux Vivre Votre Argent",
    "Ouest-France",
    "RMC",
    "Télérama",
    "Europe 1",
}

# Normalisation des types Excel → JSON
TYPE_MAP = {
    "TV/radio":   "Radio/TV",
    "Non médias": "Non médias",
    "Non m\xe9dias": "Non médias",
}

VALID_TYPES = {
    "Web", "Print", "Print & Web", "Radio/TV", "Podcast",
    "Newsletter", "Web radio", "Youtube", "Non médias",
}


def normalize_type(type_val):
    if not type_val:
        return "Web"
    t = str(type_val).strip()
    if t in TYPE_MAP:
        return TYPE_MAP[t]
    if t in VALID_TYPES:
        return t
    return "Web"


def normalize_media(media_str):
    """Normalise les noms de médias avec espaces parasites."""
    m = media_str.strip()
    remap = {
        "Bati Info": "Batinfo",
        "Se loger":  "Se Loger",
        "Plusieurs sites web locaux : Orleans info, Tarbes actu": "Orléans info, Tarbes actu (sites locaux)",
    }
    return remap.get(m, m)


# ─── PARSING ──────────────────────────────────────────────────────────────────

def parse_retombees(ws):
    retombees = []
    counter = 1

    for row_idx in range(2, ws.max_row + 1):
        media_val = ws.cell(row=row_idx, column=1).value
        type_val  = ws.cell(row=row_idx, column=2).value
        date_val  = ws.cell(row=row_idx, column=3).value
        titre_val = ws.cell(row=row_idx, column=4).value
        url_val   = ws.cell(row=row_idx, column=5).value

        if media_val is None and date_val is None:
            continue

        # Date
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        elif date_val:
            date_str = str(date_val)
        else:
            continue

        media_str = normalize_media(str(media_val).strip()) if media_val else ""
        titre_str = str(titre_val).strip().replace("\xa0", " ") if titre_val else f"Article – {media_str}"
        url_str   = str(url_val).strip() if url_val else ""
        if url_str.lower() in ("none", "nan", "-"):
            url_str = ""

        media_type = normalize_type(type_val)
        logo_media = LOGOS_MEDIAS.get(media_str, "")

        retombees.append({
            "id":          counter,
            "titre":       titre_str,
            "media":       media_str,
            "logoMedia":   logo_media,
            "date":        date_str,
            "type":        media_type,
            "format":      "",
            "url":         url_str,
            "journaliste": "",
            "pdfPath":     "",
        })
        counter += 1

    return retombees


def compute_kpis(retombees):
    types      = set(r["type"] for r in retombees)
    tv_radio   = sum(1 for r in retombees if r["type"] == "Radio/TV")
    medias_cles = set(r["media"] for r in retombees if r["media"] in MEDIAS_CLES_REF)
    return {
        "retombees":   str(len(retombees)),
        "tvRadio":     str(tv_radio),
        "typesPresse": str(len(types)),
        "mediasCles":  str(len(medias_cles)),
    }


def compute_periode(retombees):
    dates = []
    for r in retombees:
        try:
            dates.append(datetime.strptime(r["date"], "%Y-%m-%d"))
        except Exception:
            pass
    if not dates:
        return str(datetime.now().year)
    min_d = min(dates)
    max_d = max(dates)
    mois_fr = ["janv.", "févr.", "mars", "avr.", "mai", "juin",
               "juil.", "août", "sept.", "oct.", "nov.", "déc."]
    if min_d.year == max_d.year:
        return f"{mois_fr[min_d.month-1]} – {mois_fr[max_d.month-1]} {max_d.year}"
    return f"{mois_fr[min_d.month-1]} {min_d.year} – {mois_fr[max_d.month-1]} {max_d.year}"


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"Fichier Excel introuvable : {EXCEL_FILE}")
        return

    print(f"Lecture de {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Feuil1"]

    retombees = parse_retombees(ws)
    kpis      = compute_kpis(retombees)
    periode   = compute_periode(retombees)

    print(f"{len(retombees)} retombees trouvees")
    print(f"KPIs: {kpis}")

    os.makedirs(os.path.join("public", "data"), exist_ok=True)

    output = {
        "derniere_mise_a_jour": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "client": {
            "nom":             "Béa",
            "logo":            "/logo-bea.png",
            "couleurPrimaire": "#1A1A2E",
            "couleurAccent":   "#E94560",
            "periode":         periode,
            "kpis":            kpis,
        },
        "retombees": retombees,
    }

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"Fichier genere : {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
